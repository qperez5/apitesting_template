from random import shuffle
import openpyxl

path = '/home/qperez/Downloads/ejemplo_lista.xlsx'
work_book = openpyxl.load_workbook(path)
sheet = work_book['Sheet1']
rows = sheet.max_row
cols = sheet.max_column
cell1 = sheet['A2']
x3 = sheet.cell(row=rows-1, column=cols-1)


print(cell1.value)
print('los valores de celda 3: ', x3.value)


